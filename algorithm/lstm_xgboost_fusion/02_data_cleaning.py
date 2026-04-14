from __future__ import annotations

# [02.1] ===== 基础库导入 =====
import csv
import json
import math
import statistics
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import numpy as np
from openpyxl import Workbook, load_workbook
from collections import defaultdict

# [02.2] ===== 来自配置模块的常量 =====
import importlib.util

_cfg_spec = importlib.util.spec_from_file_location("cfg", Path(__file__).with_name("01_config.py"))
cfg = importlib.util.module_from_spec(_cfg_spec)
assert _cfg_spec and _cfg_spec.loader
_cfg_spec.loader.exec_module(cfg)


# [02.3] ===== 通用清洗函数 =====
def excel_serial_to_date(value: Any) -> date | None:
    """将 Excel 序列日期/日期对象统一转换为 `date`。"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    return None


def safe_float(value: Any) -> float | None:
    """鲁棒数值解析：支持数值/字符串，自动处理空值、NA、百分号等。"""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(float(value)):
            return None
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace("%", "")
        if cleaned in {"", "-", "—", "NA", "None", "nan"}:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def mean_or_none(values: list[float | None]) -> float | None:
    valid = [float(v) for v in values if v is not None]
    if not valid:
        return None
    return float(np.mean(valid))


def normalize_stage_label(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip().upper().replace(" ", "")
    if not text:
        return None
    text = text.replace("－", "-")
    if text == "VT":
        return text
    if text.startswith("V") and len(text) >= 2:
        text = "V" + text[1:]
    return text


def stage_to_code(stage_label: str | None) -> float | None:
    if stage_label is None:
        return None
    cleaned = normalize_stage_label(stage_label)
    if cleaned in cfg.STAGE_CODE_BASE:
        return cfg.STAGE_CODE_BASE[cleaned]
    if "-" in cleaned:
        parts = [part for part in cleaned.split("-") if part]
        values = [cfg.STAGE_CODE_BASE[part] for part in parts if part in cfg.STAGE_CODE_BASE]
        if values:
            return float(np.mean(values))
    return None


def mode_or_first(values: list[str]) -> str | None:
    if not values:
        return None
    cleaned = [v for v in values if v]
    if not cleaned:
        return None
    modes = statistics.multimode(cleaned)
    return modes[0] if modes else cleaned[0]


# [02.4] ===== 输入读取与解压 =====
def unzip_inputs() -> tuple[Path, Path]:
    """
    输入文件定位函数（直接读取 data 目录下文件，不生成 unzipped 目录）。
    """

    # [02.4.1] ===== 直接读取 data =====
    survey_path = cfg.DATA_DIR / cfg.SURVEY_FILENAME
    weather_path = cfg.DATA_DIR / cfg.WEATHER_FILENAME

    if not survey_path.exists():
        raise FileNotFoundError(f"未找到调查数据文件: {survey_path}")
    if not weather_path.exists():
        raise FileNotFoundError(f"未找到气象数据文件: {weather_path}")

    return survey_path, weather_path


def read_docx_text(path: Path) -> str:
    """从 docx 中提取正文段落文本（读取 word/document.xml）。"""
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ET.fromstring(xml)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs = []
    for node in root.findall(".//w:p", ns):
        text = "".join((part.text or "") for part in node.findall(".//w:t", ns)).strip()
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


# [02.5] ===== 气象数据处理 =====
def compute_relative_humidity(specific_humidity: float, temp_c: float, pressure_kpa: float) -> float:
    saturation_vapor_pressure = 0.6108 * math.exp((17.27 * temp_c) / (temp_c + 237.3))
    actual_vapor_pressure = (specific_humidity * pressure_kpa) / (0.622 + 0.378 * specific_humidity)
    if saturation_vapor_pressure <= 1e-6:
        return 0.0
    return float(np.clip(100.0 * actual_vapor_pressure / saturation_vapor_pressure, 0.0, 100.0))


def compute_rain_streaks(precip_series: np.ndarray, rain_threshold: float = 0.1) -> tuple[np.ndarray, np.ndarray]:
    rainy_streak = np.zeros(len(precip_series), dtype=np.float32)
    rain_gap = np.zeros(len(precip_series), dtype=np.float32)
    for idx, value in enumerate(precip_series):
        if value > rain_threshold:
            rainy_streak[idx] = 1.0 if idx == 0 else rainy_streak[idx - 1] + 1.0
            rain_gap[idx] = 0.0
        else:
            rainy_streak[idx] = 0.0
            rain_gap[idx] = 1.0 if idx == 0 else rain_gap[idx - 1] + 1.0
    return rainy_streak, rain_gap

# =====新增一个通用“连续天数”计算函数=====
def compute_boolean_streaks(flag_series: np.ndarray) -> np.ndarray:
    """
    对布尔条件序列计算“截至当天的连续满足天数”。
    例如 [False, True, True, False, True] -> [0,1,2,0,1]
    """
    streak = np.zeros(len(flag_series), dtype=np.float32)
    for idx, flag in enumerate(flag_series):
        if bool(flag):
            streak[idx] = 1.0 if idx == 0 else streak[idx - 1] + 1.0
        else:
            streak[idx] = 0.0
    return streak


def read_station_metadata(weather_path: Path) -> tuple[dict[int, dict[str, Any]], list[str]]:
    workbook = load_workbook(weather_path, data_only=True, read_only=True)
    sheets = workbook.sheetnames
    metadata_sheet = workbook[sheets[0]]
    metadata: dict[int, dict[str, Any]] = {}
    for row in metadata_sheet.iter_rows(min_row=2, values_only=True):
        site_id = safe_float(row[0])
        if site_id is None:
            continue
        site_id_int = int(site_id)
        metadata[site_id_int] = {
            "site_id": site_id_int,
            "province": row[1],
            "city": row[2],
            "canonical_name": row[3],
            "altitude": safe_float(row[9]),
            "lat": safe_float(row[10]),
            "lon": safe_float(row[11]),
            "download_start": excel_serial_to_date(row[12]),
            "first_survey": excel_serial_to_date(row[13]),
            "last_survey": excel_serial_to_date(row[14]),
            "weather_sheet": sheets[site_id_int],
        }
    return metadata, sheets[1:]


def read_weather_series(weather_path: Path, metadata: dict[int, dict[str, Any]]) -> dict[int, dict[str, Any]]:
    workbook = load_workbook(weather_path, data_only=True, read_only=True)
    weather_by_site: dict[int, dict[str, Any]] = {}
    for site_id, meta in metadata.items():
        sheet = workbook[meta["weather_sheet"]]
        rows = list(sheet.iter_rows(values_only=True))
        header = rows[0]
        index = {name: i for i, name in enumerate(header)}
        records = []
        for row in rows[1:]:
            obs_date = excel_serial_to_date(row[index["日期"]])
            if obs_date is None:
                continue
            record = {"date": obs_date}
            valid = True
            for feature_name, (column_name, transform) in cfg.WEATHER_COLUMN_MAP.items():
                raw_value = safe_float(row[index[column_name]])
                if raw_value is None:
                    valid = False
                    break
                record[feature_name] = float(transform(raw_value))
            if not valid:
                continue
            record["relative_humidity"] = compute_relative_humidity(
                record["humidity_avg"],
                record["temp_avg_c"],
                record["pressure_kpa"],
            )
            record["relative_humidity_max"] = compute_relative_humidity(
                record["humidity_max"],
                record["temp_min_c"],
                record["pressure_max_kpa"],
            )
            record["relative_humidity_min"] = compute_relative_humidity(
                record["humidity_min"],
                record["temp_max_c"],
                record["pressure_min_kpa"],
            )
            records.append(record)
        records.sort(key=lambda item: item["date"])
        dates = [item["date"] for item in records]
        date_index = {value: idx for idx, value in enumerate(dates)}
        matrix = np.array([[row[name] for name in cfg.SEQ_FEATURES] for row in records], dtype=np.float32)
        aux_feature_names = [name for name in records[0].keys() if name != "date"]
        aux_arrays = {name: np.array([row[name] for row in records], dtype=np.float32) for name in aux_feature_names}
        rainy_streak, rain_gap = compute_rain_streaks(aux_arrays["precip_sum"])
        aux_arrays["rainy_streak_days"] = rainy_streak
        aux_arrays["rain_gap_days"] = rain_gap

        # ===============================
        # 温度绝对阈值 + 湿度分档 + 事件标记
        # ===============================
        temp_series = aux_arrays["temp_avg_c"]
        rh_series = aux_arrays["relative_humidity"]

        temp_low_threshold = cfg.TEMP_LOW_THRESHOLD
        temp_high_threshold = cfg.TEMP_HIGH_THRESHOLD

        low_temp_flag = temp_series < cfg.TEMP_LOW_THRESHOLD
        optimal_temp_flag = (temp_series >= cfg.TEMP_OPTIMAL_LOW) & (temp_series <= cfg.TEMP_OPTIMAL_HIGH)
        high_temp_flag = temp_series > cfg.TEMP_HIGH_THRESHOLD

        high_humidity_flag = rh_series >= cfg.HIGH_HUMIDITY_THRESHOLD
        medium_high_humidity_flag = (rh_series >= cfg.MEDIUM_HIGH_HUMIDITY_THRESHOLD) & (rh_series < cfg.HIGH_HUMIDITY_THRESHOLD)

        heavy_rain_flag = aux_arrays["precip_sum"] >= cfg.HEAVY_RAIN_THRESHOLD
        weak_wind_flag = aux_arrays["wind_avg"] <= cfg.WEAK_WIND_THRESHOLD
        low_radiation_flag = aux_arrays["radiation_avg"] <= cfg.LOW_RADIATION_THRESHOLD

        # ===== 单条件连续天数 =====
        aux_arrays["temp_high_streak_days"] = compute_boolean_streaks(high_temp_flag)
        aux_arrays["temp_low_streak_days"] = compute_boolean_streaks(low_temp_flag)
        aux_arrays["temp_optimal_streak_days"] = compute_boolean_streaks(optimal_temp_flag)

        aux_arrays["high_humidity_flag"] = high_humidity_flag.astype(np.float32)
        aux_arrays["medium_high_humidity_flag"] = medium_high_humidity_flag.astype(np.float32)
        aux_arrays["heavy_rain_flag"] = heavy_rain_flag.astype(np.float32)
        aux_arrays["weak_wind_flag"] = weak_wind_flag.astype(np.float32)
        aux_arrays["low_radiation_flag"] = low_radiation_flag.astype(np.float32)

        aux_arrays["high_humidity_streak_days"] = compute_boolean_streaks(high_humidity_flag)
        aux_arrays["medium_high_humidity_streak_days"] = compute_boolean_streaks(medium_high_humidity_flag)
        aux_arrays["heavy_rain_streak_days"] = compute_boolean_streaks(heavy_rain_flag)
        aux_arrays["weak_wind_streak_days"] = compute_boolean_streaks(weak_wind_flag)
        aux_arrays["low_radiation_streak_days"] = compute_boolean_streaks(low_radiation_flag)

        # ===== 温差 / 湿度日较差 =====
        aux_arrays["temp_range_24h_c"] = aux_arrays["temp_max_c"] - aux_arrays["temp_min_c"]
        aux_arrays["humidity_range_daily"] = aux_arrays["relative_humidity_max"] - aux_arrays["relative_humidity_min"]

        # ===== 复合连续天数（高温高湿 / 适温高湿 / 弱风高湿）=====
        hot_humid_flag = high_temp_flag & high_humidity_flag
        optimal_temp_humid_flag = optimal_temp_flag & high_humidity_flag
        weak_wind_humid_flag = weak_wind_flag & high_humidity_flag

        aux_arrays["hot_humid_streak_days"] = compute_boolean_streaks(hot_humid_flag)
        aux_arrays["optimal_temp_humid_streak_days"] = compute_boolean_streaks(optimal_temp_humid_flag)
        aux_arrays["weak_wind_humid_streak_days"] = compute_boolean_streaks(weak_wind_humid_flag)

        weather_by_site[site_id] = {
            "meta": meta,
            "records": records,
            "dates": dates,
            "date_index": date_index,
            "matrix": matrix,
            "arrays": aux_arrays,
            "min_date": dates[0],
            "max_date": dates[-1],
            "temp_thresholds": {
                "low": temp_low_threshold,
                "high": temp_high_threshold,
            },
        }

        #  # ===============================
        # # 新增：站点温度阈值和事件标记
        # # ===============================
        # temp_series = aux_arrays["temp_avg_c"]
        # temp_low_thr = float(np.quantile(temp_series, cfg.TEMP_LOW_Q))
        # temp_high_thr = float(np.quantile(temp_series, cfg.TEMP_HIGH_Q))
        # temp_high_flag = temp_series >= temp_high_thr
        # temp_low_flag = temp_series <= temp_low_thr
        # temp_optimal_flag = (temp_series > temp_low_thr) & (temp_series < temp_high_thr)
        # high_humidity_flag = aux_arrays["relative_humidity"] >= cfg.HIGH_HUMIDITY_THRESHOLD
        # heavy_rain_flag = aux_arrays["precip_sum"] >= cfg.HEAVY_RAIN_THRESHOLD
        # aux_arrays["temp_high_streak_days"] = compute_boolean_streaks(temp_high_flag)
        # aux_arrays["temp_low_streak_days"] = compute_boolean_streaks(temp_low_flag)
        # aux_arrays["temp_optimal_streak_days"] = compute_boolean_streaks(temp_optimal_flag)
        # aux_arrays["high_humidity_streak_days"] = compute_boolean_streaks(high_humidity_flag)
        # aux_arrays["heavy_rain_streak_days"] = compute_boolean_streaks(heavy_rain_flag)
        # aux_arrays["high_humidity_flag"] = high_humidity_flag.astype(np.float32)
        # aux_arrays["heavy_rain_flag"] = heavy_rain_flag.astype(np.float32)

        # # ===== 温差 / 湿度日较差 =====
        # aux_arrays["temp_range_24h_c"] = aux_arrays["temp_max_c"] - aux_arrays["temp_min_c"]
        # aux_arrays["humidity_range_daily"] = aux_arrays["relative_humidity_max"] - aux_arrays["relative_humidity_min"]

        # # ===== 站点温度阈值 =====
        # temp_low_threshold = float(np.quantile(aux_arrays["temp_avg_c"], cfg.TEMP_LOW_Q))
        # temp_high_threshold = float(np.quantile(aux_arrays["temp_avg_c"], cfg.TEMP_HIGH_Q))

        # high_temp_flag = aux_arrays["temp_avg_c"] >= temp_high_threshold
        # low_temp_flag = aux_arrays["temp_avg_c"] <= temp_low_threshold
        # optimal_temp_flag = (aux_arrays["temp_avg_c"] > temp_low_threshold) & (aux_arrays["temp_avg_c"] < temp_high_threshold)

        # aux_arrays["temp_high_streak_days"] = compute_boolean_streaks(high_temp_flag)
        # aux_arrays["temp_low_streak_days"] = compute_boolean_streaks(low_temp_flag)
        # aux_arrays["temp_optimal_streak_days"] = compute_boolean_streaks(optimal_temp_flag)

        # # ===== 高湿 / 强降雨 / 弱风 / 寡照 =====
        # high_humidity_flag = aux_arrays["relative_humidity"] >= cfg.HIGH_HUMIDITY_THRESHOLD
        # heavy_rain_flag = aux_arrays["precip_sum"] >= cfg.HEAVY_RAIN_THRESHOLD
        # weak_wind_flag = aux_arrays["wind_avg"] <= cfg.WEAK_WIND_THRESHOLD
        # low_radiation_flag = aux_arrays["radiation_avg"] <= cfg.LOW_RADIATION_THRESHOLD

        # aux_arrays["high_humidity_flag"] = high_humidity_flag.astype(np.float32)
        # aux_arrays["heavy_rain_flag"] = heavy_rain_flag.astype(np.float32)
        # aux_arrays["weak_wind_flag"] = weak_wind_flag.astype(np.float32)
        # aux_arrays["low_radiation_flag"] = low_radiation_flag.astype(np.float32)

        # aux_arrays["high_humidity_streak_days"] = compute_boolean_streaks(high_humidity_flag)
        # aux_arrays["heavy_rain_streak_days"] = compute_boolean_streaks(heavy_rain_flag)
        # aux_arrays["weak_wind_streak_days"] = compute_boolean_streaks(weak_wind_flag)
        # aux_arrays["low_radiation_streak_days"] = compute_boolean_streaks(low_radiation_flag)

        # # ===== 复合连续天数 =====
        # hot_humid_flag = high_temp_flag & high_humidity_flag
        # optimal_temp_humid_flag = optimal_temp_flag & high_humidity_flag
        # weak_wind_humid_flag = weak_wind_flag & high_humidity_flag

        # aux_arrays["hot_humid_streak_days"] = compute_boolean_streaks(hot_humid_flag)
        # aux_arrays["optimal_temp_humid_streak_days"] = compute_boolean_streaks(optimal_temp_humid_flag)
        # aux_arrays["weak_wind_humid_streak_days"] = compute_boolean_streaks(weak_wind_humid_flag)
        # weather_by_site[site_id] = {
        #     "meta": meta,
        #     "records": records,
        #     "dates": dates,
        #     "date_index": date_index,
        #     "matrix": matrix,
        #     "arrays": aux_arrays,
        #     "min_date": dates[0],
        #     "max_date": dates[-1],
        #     "temp_thresholds": {
        #         "low": temp_low_threshold,
        #         "high": temp_high_threshold,
        #     },
        # }
    return weather_by_site


# [02.6] ===== 调查数据聚合 =====
def adjust_survey_date(raw_date: date, weather_site: dict[str, Any]) -> tuple[date, int]:
    min_date = weather_site["min_date"] - timedelta(days=3)
    max_date = weather_site["max_date"] + timedelta(days=3)
    if min_date <= raw_date <= max_date:
        return raw_date, 0
    shifted = raw_date - timedelta(days=365)
    if min_date <= shifted <= max_date:
        return shifted, -365
    return raw_date, 0


def read_and_aggregate_survey(survey_path: Path, weather_by_site: dict[int, dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = load_workbook(survey_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    header = rows[0]
    index = {name: i for i, name in enumerate(header)}
    groups: dict[tuple[int, date], dict[str, Any]] = {}
    quality_rows: list[dict[str, Any]] = []
    # for row in rows[1:]:
    for raw_idx, row in enumerate(rows[1:], start=2):
        site_id_raw = safe_float(row[index["序号"]])
        if site_id_raw is None:
            continue
        site_id = int(site_id_raw)
        raw_date = excel_serial_to_date(row[index["时间"]])
        if raw_date is None:
            continue
        adjusted_date, shift_days = adjust_survey_date(raw_date, weather_by_site[site_id])
        key = (site_id, adjusted_date, raw_idx)
        if key not in groups:
            groups[key] = {
                "site_id": site_id,
                "date": adjusted_date,
                "record_id": raw_idx,
                "raw_dates": [],
                "shift_days": [],
                "site_names": [],
                "varieties": set(),
                "stage_labels": [],
                "stage_codes": [],
                "gray_resistance_values": [],
                "blight_resistance_values": [],
                "white_resistance_values": [],
                "gray_incidence_values": [],
                "gray_index_values": [],
                "blight_incidence_values": [],
                "blight_index_values": [],
                "white_incidence_values": [],
                "white_index_values": [],
            }
        state = groups[key]
        state["raw_dates"].append(raw_date.isoformat())
        state["shift_days"].append(shift_days)
        if row[index["地点"]] is not None:
            state["site_names"].append(str(row[index["地点"]]).strip())
        if row[index["品种"]] is not None and str(row[index["品种"]]).strip():
            state["varieties"].add(str(row[index["品种"]]).strip())
        stage_label = normalize_stage_label(row[index["生育期"]])
        stage_code = stage_to_code(stage_label)
        if stage_label is not None:
            state["stage_labels"].append(stage_label)
        if stage_code is not None:
            state["stage_codes"].append(stage_code)
        for field, column in [
            ("gray_resistance_values", "灰斑病抗性"),
            ("blight_resistance_values", "大斑病抗性"),
            ("white_resistance_values", "白斑病抗性"),
            ("gray_incidence_values", "灰斑病发病株率"),
            ("gray_index_values", "灰斑病病情指数"),
            ("blight_incidence_values", "大斑病发病株率"),
            ("blight_index_values", "大斑病病情指数"),
            ("white_incidence_values", "白斑病发病株率"),
            ("white_index_values", "白斑病病情指数"),
        ]:
            parsed = safe_float(row[index[column]])
            if parsed is not None:
                state[field].append(parsed)
        quality_rows.append(
            {
                "site_id": site_id,
                "raw_date": raw_date.isoformat(),
                "adjusted_date": adjusted_date.isoformat(),
                "shift_days": shift_days,
                "raw_site_name": str(row[index["地点"]]).strip() if row[index["地点"]] is not None else "",
                "variety": str(row[index["品种"]]).strip() if row[index["品种"]] is not None else "",
            }
        )
    panel_rows = []
    for (site_id, obs_date, record_id), state in sorted(
        groups.items(), 
        key=lambda item: (item[0][0], item[0][1], item[0][2])):
        panel_rows.append(
            {
                "site_id": site_id,
                "site_name": mode_or_first(state["site_names"]) or f"site_{site_id}",
                "date": obs_date,
                "date_str": obs_date.isoformat(),

                "record_id": record_id,

                "sample_count": len(state["site_names"]) if state["site_names"] else 0,
                "varieties": ";".join(sorted(state["varieties"])),

                "stage_label": mode_or_first(state["stage_labels"]),
                "stage_code": mean_or_none(state["stage_codes"]),
                "gray_resistance": mean_or_none(state["gray_resistance_values"]),
                "blight_resistance": mean_or_none(state["blight_resistance_values"]),
                "white_resistance": mean_or_none(state["white_resistance_values"]),
                "gray_incidence": mean_or_none(state["gray_incidence_values"]),
                "gray_index": mean_or_none(state["gray_index_values"]),
                "blight_incidence": mean_or_none(state["blight_incidence_values"]),
                "blight_index": mean_or_none(state["blight_index_values"]),
                "white_incidence": mean_or_none(state["white_incidence_values"]),
                "white_index": mean_or_none(state["white_index_values"]),
                "shift_days_total": int(sum(state["shift_days"])),
                "shifted_row_count": int(sum(1 for value in state["shift_days"] if value != 0)),
            }
        )

        same_day_counter = defaultdict(int)
        for row in sorted(panel_rows, key=lambda x: (x["site_id"], x["date"], x["record_id"])):
            key2 = (row["site_id"], row["date"])
            same_day_counter[key2] += 1
            row["replicate_id_same_day"] = same_day_counter[key2]
            
    return panel_rows, quality_rows


# [02.7] ===== 导出函数 =====
def write_csv(path: Path, rows: list[dict[str, Any]], field_order: list[str] | None = None) -> None:
    if not rows:
        return
    if field_order is None:
        field_order = list(rows[0].keys())
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=field_order)
        writer.writeheader()
        for row in rows:
            sanitized = {}
            for key in field_order:
                value = row.get(key)
                if isinstance(value, date):
                    sanitized[key] = value.isoformat()
                elif isinstance(value, np.ndarray):
                    sanitized[key] = json.dumps(value.tolist(), ensure_ascii=False)
                else:
                    sanitized[key] = value
            writer.writerow(sanitized)


def write_workbook(path: Path, sheets: list[tuple[str, list[dict[str, Any]]]]) -> None:
    workbook = Workbook()
    first = True
    for sheet_name, rows in sheets:
        worksheet = workbook.active if first else workbook.create_sheet()
        first = False
        worksheet.title = sheet_name[:31]
        if not rows:
            continue
        headers = list(rows[0].keys())
        worksheet.append(headers)
        for row in rows:
            values = []
            for header in headers:
                value = row.get(header)
                if isinstance(value, date):
                    values.append(value.isoformat())
                elif isinstance(value, np.ndarray):
                    values.append(json.dumps(value.tolist(), ensure_ascii=False))
                else:
                    values.append(value)
            worksheet.append(values)
        worksheet.freeze_panes = "A2"
    workbook.save(path)
