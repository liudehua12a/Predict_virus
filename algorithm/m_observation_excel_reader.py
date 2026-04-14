from __future__ import annotations

from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import k_weather_data_storage as storage


HEADER_ALIASES = {
    "date": "date",
    "site_name": "site_name",
    "batch_name": "batch_name",
    "crop_variety": "crop_variety",
    "growth_stage": "growth_stage",
    "gray_incidence": "gray_incidence",
    "gray_index": "gray_index",
    "blight_incidence": "blight_incidence",
    "blight_index": "blight_index",
    "white_incidence": "white_incidence",
    "white_index": "white_index",
}


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    return HEADER_ALIASES.get(text, text)


def normalize_date_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    text = str(value).strip()
    return text[:10]


def normalize_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace("%", "")
    if text == "":
        return None
    return float(text)


def get_template_sheet(workbook) -> Any:
    return workbook[workbook.sheetnames[0]]


def read_observation_excel(file_path: str | Path) -> list[dict[str, Any]]:
    """
    读取真实调查 Excel 模板：
    - 第1行：中文表头
    - 第2行：英文字段
    - 第3行开始：数据
    """
    file_path = Path(file_path)
    workbook = load_workbook(file_path, data_only=True)
    sheet = get_template_sheet(workbook)

    english_headers_raw = [cell.value for cell in sheet[2]]
    english_headers = [normalize_header(x) for x in english_headers_raw]

    rows: list[dict[str, Any]] = []

    for excel_row_no, row in enumerate(
        sheet.iter_rows(min_row=3, values_only=True),
        start=3,
    ):
        if row is None:
            continue

        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        record = {}
        for idx, field_name in enumerate(english_headers):
            if not field_name:
                continue
            value = row[idx] if idx < len(row) else None
            record[field_name] = value

        record["_source_row_no"] = excel_row_no
        record["_source_file_name"] = file_path.name
        rows.append(record)

    return rows


def map_excel_row_to_observation_record(
    excel_row: dict[str, Any],
) -> dict[str, Any]:
    """
    将 Excel 原始行转换为 disease_observation 入库结构。
    通过 site_name + batch_name 联查 site_id / batch_id。
    """
    site_name_raw = excel_row.get("site_name")
    batch_name_raw = excel_row.get("batch_name")

    if site_name_raw is None or str(site_name_raw).strip() == "":
        raise ValueError(
            f"Excel第{excel_row.get('_source_row_no')}行未填写 site_name。"
        )

    if batch_name_raw is None or str(batch_name_raw).strip() == "":
        raise ValueError(
            f"Excel第{excel_row.get('_source_row_no')}行未填写 batch_name。"
        )

    site_name = str(site_name_raw).strip()
    batch_name = str(batch_name_raw).strip()

    site_batch_row = storage.get_site_batch_by_names(
        site_name=site_name,
        batch_name=batch_name,
    )
    if not site_batch_row:
        raise ValueError(
            f"Excel第{excel_row.get('_source_row_no')}行未匹配到 site_name + batch_name："
            f"{site_name} / {batch_name}"
        )

    return {
        "site_id": int(site_batch_row["site_id"]),
        "batch_id": int(site_batch_row["batch_id"]),
        "survey_date": normalize_date_value(excel_row["date"]),
        "crop_variety": None if excel_row.get("crop_variety") is None else str(excel_row.get("crop_variety")).strip(),
        "growth_stage": None if excel_row.get("growth_stage") is None else str(excel_row.get("growth_stage")).strip(),
        "source_file_name": excel_row["_source_file_name"],
        "source_row_no": int(excel_row["_source_row_no"]),

        "gray_incidence": normalize_float(excel_row.get("gray_incidence")),
        "gray_index": normalize_float(excel_row.get("gray_index")),

        "blight_incidence": normalize_float(excel_row.get("blight_incidence")),
        "blight_index": normalize_float(excel_row.get("blight_index")),

        "white_incidence": normalize_float(excel_row.get("white_incidence")),
        "white_index": normalize_float(excel_row.get("white_index")),
    }


def read_and_map_observation_excel(file_path: str | Path) -> list[dict[str, Any]]:
    raw_rows = read_observation_excel(file_path)

    print(f"[调试] Excel读取到原始数据行数: {len(raw_rows)}")
    for idx, row in enumerate(raw_rows[:5], start=1):
        print(f"[调试] raw_rows[{idx}] =", row)

    mapped_rows = [map_excel_row_to_observation_record(row) for row in raw_rows]
    return mapped_rows


if __name__ == "__main__":
    demo_path = Path(__file__).resolve().parent / "data" / "调查数据表--模板.xlsx"
    rows = read_and_map_observation_excel(demo_path)

    print(f"读取并映射成功，共 {len(rows)} 条")
    for row in rows[:5]:
        print(row)