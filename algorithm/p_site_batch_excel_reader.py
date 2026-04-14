from __future__ import annotations

from pathlib import Path
from typing import Any
from datetime import datetime, date

from openpyxl import load_workbook


HEADER_ALIASES = {
    "province": "province",
    "city": "city",
    "site_name": "site_name",
    "lat": "lat",
    "lon": "lon",
    "elevation": "elevation",
    "batch_name": "batch_name",
    "crop_variety": "crop_variety",
    # "sowing_date": "sowing_date",
}


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    return HEADER_ALIASES.get(text, text)


def normalize_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text == "":
        return None
    return float(text)


def normalize_date_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if text == "":
        return None
    return text[:10]


def split_semicolon_values(value: Any) -> list[str]:
    """
    按中文分号/英文分号拆分，并去掉空白项。
    支持：
    - A；B；C
    - A;B;C
    """
    if value is None:
        return []

    text = str(value).strip()
    if text == "":
        return []

    text = text.replace(";", "；")
    parts = [x.strip() for x in text.split("；")]
    return [x for x in parts if x != ""]


def get_template_sheet(workbook) -> Any:
    return workbook[workbook.sheetnames[0]]


def read_site_batch_excel(file_path: str | Path) -> list[dict[str, Any]]:
    """
    读取点位+批次基础信息表：
    - 第1行：中文说明
    - 第2行：英文系统字段
    - 第3行开始：数据
    """
    file_path = Path(file_path)
    workbook = load_workbook(file_path, data_only=True)
    sheet = get_template_sheet(workbook)

    english_headers_raw = [cell.value for cell in sheet[2]]
    english_headers = [normalize_header(x) for x in english_headers_raw]

    rows: list[dict[str, Any]] = []

    for excel_row_no, row in enumerate(
        sheet.iter_rows(min_row=3, values_only=True),
        start=3,
    ):
        if row is None:
            continue

        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        record = {}
        for idx, field_name in enumerate(english_headers):
            if not field_name:
                continue
            value = row[idx] if idx < len(row) else None
            record[field_name] = value

        record["_source_row_no"] = excel_row_no
        record["_source_file_name"] = file_path.name
        rows.append(record)

    return rows


def expand_batch_rows(
    batch_name_value: Any,
    crop_variety_value: Any,
    # sowing_date_value: Any,
    source_row_no: int,
) -> list[dict[str, Any]]:
    """
    将一行中的 batch_name / crop_variety 按分号展开成多个批次。
    规则：
    1. batch_name 按 ； 拆分，必须至少有1个
    2. crop_variety 若为空，则所有批次 crop_variety = None
    3. crop_variety 若不为空，则数量必须与 batch_name 数量一致
    4. sowing_date 当前按整行共用，同一行拆出的所有批次共用一个 sowing_date
    """
    batch_names = split_semicolon_values(batch_name_value)
    crop_varieties = split_semicolon_values(crop_variety_value)
    # sowing_date = normalize_date_value(sowing_date_value)

    if len(batch_names) == 0:
        raise ValueError(f"Excel第{source_row_no}行 batch_name 为空，无法生成批次。")

    batch_rows: list[dict[str, Any]] = []

    if len(crop_varieties) == 0:
        for batch_name in batch_names:
            batch_rows.append(
                {
                    "batch_name": batch_name,
                    "batch_code": None,
                    "crop_variety": None,
                    # "sowing_date": sowing_date,
                    "survey_start_date": None,
                    "survey_end_date": None,
                }
            )
        return batch_rows

    if len(crop_varieties) != len(batch_names):
        raise ValueError(
            f"Excel第{source_row_no}行 batch_name 数量({len(batch_names)}) "
            f"与 crop_variety 数量({len(crop_varieties)}) 不一致，无法一一对应。"
        )

    for batch_name, crop_variety in zip(batch_names, crop_varieties):
        batch_rows.append(
            {
                "batch_name": batch_name,
                "batch_code": None,
                "crop_variety": crop_variety,
                # "sowing_date": sowing_date,
                "survey_start_date": None,
                "survey_end_date": None,
            }
        )

    return batch_rows


def map_excel_row_to_site_and_batches(
    excel_row: dict[str, Any],
) -> dict[str, Any]:
    """
    将 Excel 单行拆成：
    - 1 个 site_info 入库结构
    - N 个 survey_batch 入库结构
    """
    source_row_no = excel_row["_source_row_no"]

    site_row = {
        "province": None if excel_row.get("province") is None else str(excel_row["province"]).strip(),
        "city": None if excel_row.get("city") is None else str(excel_row["city"]).strip(),
        "site_name": None if excel_row.get("site_name") is None else str(excel_row["site_name"]).strip(),
        "lat": normalize_float(excel_row.get("lat")),
        "lon": normalize_float(excel_row.get("lon")),
        "elevation": normalize_float(excel_row.get("elevation")),
        "location_id": None,
    }

    batch_rows = expand_batch_rows(
        batch_name_value=excel_row.get("batch_name"),
        crop_variety_value=excel_row.get("crop_variety"),
        # sowing_date_value=excel_row.get("sowing_date"),
        source_row_no=source_row_no,
    )

    return {
        "site_row": site_row,
        "batch_rows": batch_rows,
        "_source_row_no": source_row_no,
        "_source_file_name": excel_row["_source_file_name"],
    }


def read_and_map_site_batch_excel(file_path: str | Path) -> list[dict[str, Any]]:
    raw_rows = read_site_batch_excel(file_path)
    mapped_rows = [map_excel_row_to_site_and_batches(row) for row in raw_rows]
    return mapped_rows


if __name__ == "__main__":
    demo_path = Path(__file__).resolve().parent / "data" / "点位批次基础信息表.xlsx"
    rows = read_and_map_site_batch_excel(demo_path)

    print(f"读取并映射成功，共 {len(rows)} 条点位记录")
    for row in rows[:5]:
        print(row)